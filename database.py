import config
import openpyxl
database = [[None],[None],[None],[None]]
wb = openpyxl.load_workbook(filename = 'E:/Bot/schedule.xlsx', read_only= True)
liter = None
days = None

def trySearchUserID(ID):
    if ID in database[0]:
        return True
    else:
        database[0].append(ID)
        database[1].append(0)
        database[2].append(None)
        database[3].append(None)
        return False


def checkCorrect(ID, message):
    index = database[0].index(ID)
    if (database[1][index] == 0) and ((message in config.classes) or (message in config.classes_two)):
        return True
    elif (database[1][index] == 1) and (((message in config.liters) and (database[2][index] in config.classes))or((message in config.liters_two) and (database[2][index] in config.classes_two))):
        return True
    elif (database[1][index] == 2) and (message in config.days):
        return True
    else:
        return False


#def outMessage(ID, day):
#    index = database[0].index(ID)
#    return "Расписание для " + database[2][index] + database[3][index] + ' ' + day


def processMessage (ID, message):
    trySearchUserID(ID)
    index = database[0].index(ID)
    if checkCorrect(ID, message):
        if database[1][index] == 0:
            database[2][index] = message
            database[1][index]+= 1
        elif database[1][index] == 1:
            database[3][index] = message
            database[1][index] += 1
        elif database[1][index] == 2:
            database[1][index] = 0
            return outMessage(ID, message)
    else:
        return "Некорректный ввод"


def startCommand(ID):
    if trySearchUserID(ID):
        index = database[0].index(ID)
        database[1][index] = 0


def sendCurrentInput(ID):
    index = database[0].index(ID)
    if database[1][index] == 0:
        return "Введите параллель"
    elif database[1][index] == 1:
        return "Введите литеру класса"
    elif database[1][index] == 2:
        return "Введите день недели"


def sendCurrentMarkup(ID):
    index = database[0].index(ID)
    if database[1][index] == 0:
        return config.classesMarkup
    elif database[1][index] == 1 and database[2][index] in config.classes:
        return config.litersMarkup
    elif database[1][index] == 1 and database[2][index] in config.classes_two:
        return config.liters_twoMarkup
    elif database[1][index] == 2:
        return config.daysMarkup


def outMessage(ID, day):
    index = database[0].index(ID)
    ws = wb[database[2][index]]
    for j in range(3, 8):
        if ws.cell(row=1, column=j).value == database[3][index]:
            global liter
            liter = j
            break
    for i in range(2, ws.max_row):
        if ws.cell(row=i, column=1).value == day:
            global days
            days = i
            break
    out = ''
    cycle = days
    time = ws.cell(row=cycle, column=2).value
    process = ws.cell(row=cycle, column=liter).value
    while process is not None:
        out = out + time + ': ' + process + ' \n'
        cycle += 1
        process = ws.cell(row=cycle, column=liter).value
        time = ws.cell(row=cycle, column=2).value
    return out






